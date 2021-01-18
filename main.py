# -*- coding: utf-8 -*-
import requests
from lxml import etree
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
import json
import time
import datetime
from requests import Request, Session
from requests.exceptions import ConnectionError, Timeout, TooManyRedirects
from model import Modelfromdict, ModelElement
import threading


# API 地址+文档：https://pro.coinmarketcap.com/account


# 与excel表中顺序对应,与 接口 parameters 对应（顺序不对应）
TOKENS = [
    'BTC',
    'ETH',
    'XRP',
    'BCH',
    'LTC',
    'EOS',
    'ADA',
    'XLM',
    'ETC',
    'BNB',
    'OKB',
    'HT',
    'ZEC',
    'NEO',
    'QTUM',
    'GXC',
    'ELF',
    'MIOTA',
    'WICC',
    'STORJ',
    'MDS',
    'PAI',
    'FIL'
]

# excel 位置
Excel_Position = 'block_chain_analyze.xlsx'

# 变量
usdt_rate = None
btc_coast = None
huobi_usdt_price = None
fear_greed_index = None
final_data = None


def http_get(url):
    ''' http get req '''
    headers = {
        'User-Agent': 'Mozilla/5.0 (iPhone; CPU iPhone OS 11_0 like Mac OS X) AppleWebKit/604.1.38 (KHTML, like Gecko) Version/11.0 Mobile/15A372 Safari/604.1'}
    res = requests.get(url, headers=headers, verify=False)
    res_status = res.status_code
    if res_status == 200:
        return res


def get_usdt_rate():
    ''' 获取美元汇率 '''
    # 新浪财经
    url = 'https://www.baidu.com/s?ie=utf-8&f=8&rsv_bp=1&tn=96683251_hao_pg&wd=%E7%BE%8E%E5%85%83%E6%B1%87%E7%8E%87&oq=%25E7%25BE%258E%25E5%2585%2583%25E6%25B1%2587%25E7%258E%2587%2520python%2520%25E6%258E%25A5%25E5%258F%25A3&rsv_pq=849622a40002b0a2&rsv_t=3c97aODwSAtR9tXCNOIUwwMxmD4KkKOIKMP6XkUfkYHX16WZ8o7ql3pF8Jp2OEcOzsEROyMy&rqlang=cn&rsv_enter=1&inputT=564&rsv_sug3=19&rsv_sug1=10&rsv_sug7=100&rsv_sug2=0&rsv_sug4=815&rsv_sug=2'
    res = http_get(url)
    selector = etree.HTML(res.text)
    arr = selector.xpath(
        '//*[@id="2"]/div[1]/div[1]/div[1]/div[1]/span[1]/text()')
    usdPrice = arr[0]
    print('美元价格' + usdPrice)
    global usdt_rate
    usdt_rate = usdPrice
    return usdPrice


def get_btc_coast():
    ''' 获取 BTC 挖矿成本 '''
    # 新浪财经
    url = 'https://www.trinsicoin.com/'
    res = http_get(url)
    selector = etree.HTML(res.text)
    arr = selector.xpath('//*[@id="home"]/div[2]/div/div/h1/strong/text()')
    btcCoast = arr[0]
    print('btc 挖矿成本' + btcCoast)
    global btc_coast
    btc_coast = btcCoast
    return btcCoast


def get_huobi_usdt_price():
    ''' 获取火币USDT买入价格 '''
    url = 'https://otc-api.eiijo.cn/v1/data/trade-market?country=37&currency=1&payMethod=0&currPage=1&coinId=2&tradeType=sell&blockType=general&online=1'
    res = http_get(url)
    obj = json.loads(res.text)
    huobiUsdtPrice = obj['data'][0]['price']
    print('USDT价格' + str(huobiUsdtPrice))
    global huobi_usdt_price
    huobi_usdt_price = huobiUsdtPrice
    return huobiUsdtPrice


def get_fear_greed_index():
    ''' 获取恐惧贪婪指数对象 '''
    url = 'https://api.alternative.me/fng/'
    res = http_get(url)

    resData = json.loads(res.text)
    resData = resData['data'][0]

    name = resData['value_classification']
    value = resData['value']
    # 默认绿色，在恐惧的时候显示绿色
    color = '3cb371'

    if name == 'Fear' or name == 'Extreme Fear':
        name = '恐惧'
    elif name == 'Greed' or name == 'Extreme Greed':
        name = '贪婪'
        color = 'CC0000'

    currentTime = int(resData['timestamp'])
    timeArray = time.localtime(currentTime)
    formatTime = time.strftime("%Y-%m-%d %H:%M:%S", timeArray)

    dic = {'name': name, 'value': value,
           'time': formatTime, 'color': color}
    global fear_greed_index
    fear_greed_index = value
    return dic


def write_2_excel():
    ''' 写入 excel '''
    get_final_data()

    workBook = load_workbook(Excel_Position)
    workSheet = workBook['record']

    # 价格、排名、市值等写入
    for i in range(len(final_data)):
        obj = final_data[i]
        workSheet['A{}'.format(i + 4)] = obj.tokenName
        workSheet['B{}'.format(i + 4)] = obj.rank
        workSheet['C{}'.format(
            i + 4)] = datetime.datetime.strptime(obj.lowestDate18_19, '%Y/%m/%d')
        workSheet['D{}'.format(i + 4)] = obj.lowestPrice18_19
        workSheet['E{}'.format(
            i + 4)] = datetime.datetime.strptime(obj.lowestDate18_20, '%Y/%m/%d')
        workSheet['F{}'.format(i + 4)] = obj.lowestPrice18_20
        workSheet['G{}'.format(i + 4)] = obj.usdtPrice
        workSheet['J{}'.format(i + 4)] = obj.marketCap

    # 恐惧贪婪 cell 写入
    fearGreedDic = get_fear_greed_index()
    print(fearGreedDic)
    name = fearGreedDic['name']
    value = fearGreedDic['value']
    time = fearGreedDic['time']
    cellColor = fearGreedDic['color']
    fearGreedCell = workSheet['A2']
    timeCell = workSheet['A1']

    fearGreedCell.fill = PatternFill(
        start_color=cellColor, end_color=cellColor, fill_type="solid")
    timeCell.fill = PatternFill(
        start_color=cellColor, end_color=cellColor, fill_type="solid")
    fearGreedCell.value = "{0} : {1}".format(name, value)
    timeCell.value = time

    # 美元价格写入
    workSheet['K4'] = str(get_usdt_rate())

    # 火币 USDT 价格 写入
    workSheet['K6'] = get_huobi_usdt_price()

    # BTC 挖矿成本写入
    workSheet['K2'] = get_btc_coast()

    workBook.save(Excel_Position)


def read_net_data(tokens=None):
    ''' 获取 token 价格 '''
    url = 'https://pro-api.coinmarketcap.com/v1/cryptocurrency/quotes/latest'

    parameters = {
        'symbol': 'BTC,ETH,XRP,BCH,LTC,EOS,BNB,ADA,XLM,NEO,MIOTA,ZEC,QTUM,HT,GXC,PAI,STORJ,CTXC,CMT,TNB,MFT,DX,DACC,RUFF,MDS,ELF,WICC,OKB,ETC,TRX,FIL',
    }
    if tokens is not None:
        parameters = tokens

    headers = {
        'Accepts': 'application/json',
        'X-CMC_PRO_API_KEY': 'af108ddb-cc3a-429e-b408-087dc8a71a11',
    }

    session = Session()
    session.headers.update(headers)

    result = []
    try:
        response = session.get(url, params=parameters)
        # print(response.text)
        data = json.loads(response.text)

        for obj in TOKENS:
            dic = {'tokenName': obj, 'usdtPrice': data['data'][obj]['quote']['USD']['price'],
                   'rank': data['data'][obj]['cmc_rank'], 'marketCap': data['data'][obj]['quote']['USD']['market_cap']}
            result.append(dic)

        return result

    except (ConnectionError, Timeout, TooManyRedirects) as e:
        print(e)


def read_local_data(tokens=None):
    ''' 读取本地数据 '''
    with open("data.json") as json_file:
        data = json.load(json_file)
    return data


def get_final_data(tokens=None):
    ''' 获取合并本地和远程的数据 '''
    netData = read_net_data(tokens=tokens)
    lodcalData = read_local_data(tokens=tokens)

    result = []
    for net_dic in netData:
        for local_dic in lodcalData:
            if net_dic["tokenName"] == local_dic["tokenName"]:
                local_dic["usdtPrice"] = net_dic["usdtPrice"]
                local_dic["rank"] = net_dic["rank"]
                local_dic["marketCap"] = net_dic["marketCap"]
                result.append(local_dic)

    print('😀😀😀 ============== merge data（json 格式） ============== 😀😀😀')
    print(json.dumps(result))

    list_data = Modelfromdict(result)
    print('😀😀😀 ============== list_data ============== 😀😀😀')
    print(list_data)

    result = sorted(list_data, key=lambda x: x.marketCap)
    global final_data
    final_data = result
    return result


def format_response_data(req=None):
    tokes_base_info_thread = threading.Thread(
        target=get_final_data, args=(req,))
    fear_greed_index_thread = threading.Thread(target=get_fear_greed_index)
    huobi_usdt_price_thread = threading.Thread(target=get_huobi_usdt_price)
    btc_coast_thread = threading.Thread(target=get_btc_coast)
    usdt_rate_thread = threading.Thread(target=get_usdt_rate)

    tokes_base_info_thread.start()
    fear_greed_index_thread.start()
    huobi_usdt_price_thread.start()
    btc_coast_thread.start()
    usdt_rate_thread.start()

    tokes_base_info_thread.join()
    fear_greed_index_thread.join()
    huobi_usdt_price_thread.join()
    btc_coast_thread.join()
    usdt_rate_thread.join()

    data = {
        'tokensInfo': final_data,
        'attachData': {
            'fear_greed_index': fear_greed_index,
            'huobi_usdt_price': huobi_usdt_price,
            'btc_coast': btc_coast,
            'usdt_rate': usdt_rate
        }
    }
    print(data)
    return data


def main():
    write_2_excel()


if __name__ == '__main__':
    main()
